"""Server-side table export to XLSX and CSV formats.

Uses openpyxl for XLSX (already a project dependency) and csv for CSV.
Supports both single-sheet (legacy) and multi-sheet specs.
"""

import csv
import io
import logging
from typing import Any

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Format detection helpers
# ---------------------------------------------------------------------------

def _is_multi_sheet(spec: dict) -> bool:
    """Return True if the spec uses the multi-sheet format (has ``sheets`` key)."""
    return "sheets" in spec and isinstance(spec["sheets"], list)


def _normalize_spec(spec: dict) -> tuple[list[str], list[list[Any]]]:
    """Normalize a *single-sheet* table spec into (column_labels, rows).

    Handles both flat format ``{columns: [...], rows: [[...]]}`` and
    object-row format ``{columns: [{key, label}], rows: [{key: val}]}``.

    Returns:
        Tuple of (column_labels, row_data) where row_data is a list of
        flat lists (one value per column).
    """
    raw_cols = spec.get("columns", [])
    raw_rows = spec.get("rows", [])

    # Build column labels and keys
    col_labels: list[str] = []
    col_keys: list[str] = []
    for c in raw_cols:
        if isinstance(c, dict):
            col_labels.append(str(c.get("label", c.get("key", ""))))
            col_keys.append(str(c.get("key", c.get("label", ""))))
        else:
            col_labels.append(str(c))
            col_keys.append(str(c))

    # Normalize rows
    rows: list[list[Any]] = []
    for row in raw_rows:
        if isinstance(row, dict):
            rows.append([row.get(k, "") for k in col_keys])
        elif isinstance(row, (list, tuple)):
            rows.append(list(row))
        else:
            rows.append([row])

    return col_labels, rows


def _normalize_multi_sheet(
    spec: dict,
) -> list[tuple[str, list[dict], list[list[Any]], dict]]:
    """Normalize a multi-sheet spec.

    Returns:
        List of ``(sheet_name, column_defs, normalized_rows, sheet_opts)``
        where *column_defs* keeps the full dicts (with ``width``, ``format``),
        *normalized_rows* is a list of flat lists, and *sheet_opts* contains
        per-sheet options (``frozen_rows``, ``auto_filter``).
    """
    result: list[tuple[str, list[dict], list[list[Any]], dict]] = []
    for sheet in spec["sheets"]:
        name = str(sheet.get("name", "Sheet"))[:31]
        raw_cols: list = sheet.get("columns", [])
        raw_rows: list = sheet.get("rows", [])

        # Normalize column defs to dicts
        col_defs: list[dict] = []
        col_keys: list[str] = []
        for c in raw_cols:
            if isinstance(c, dict):
                col_defs.append(c)
                col_keys.append(str(c.get("key", c.get("label", ""))))
            else:
                col_defs.append({"key": str(c), "label": str(c)})
                col_keys.append(str(c))

        # Normalize rows – also collect per-row _style
        rows: list[list[Any]] = []
        row_styles: list[dict | None] = []
        for row in raw_rows:
            if isinstance(row, dict):
                style = row.get("_style")
                rows.append([row.get(k, "") for k in col_keys])
                row_styles.append(style if isinstance(style, dict) else None)
            elif isinstance(row, (list, tuple)):
                rows.append(list(row))
                row_styles.append(None)
            else:
                rows.append([row])
                row_styles.append(None)

        sheet_opts: dict = {
            "frozen_rows": sheet.get("frozen_rows"),
            "auto_filter": sheet.get("auto_filter", False),
            "row_styles": row_styles,
        }
        result.append((name, col_defs, rows, sheet_opts))
    return result


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

def export_xlsx(spec: dict) -> bytes:
    """Generate XLSX from a table spec.

    Supports both the legacy single-sheet format and the multi-sheet format
    (detected by the presence of a ``sheets`` key).

    Args:
        spec: Table spec with ``columns``/``rows`` or ``sheets``.

    Returns:
        XLSX file content as bytes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Default styling shared by both paths
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1D459F", end_color="1D459F", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    default_cell_font = Font(name="Calibri", size=11)

    if _is_multi_sheet(spec):
        sheets_data = _normalize_multi_sheet(spec)

        for sheet_idx, (name, col_defs, rows, opts) in enumerate(sheets_data):
            if sheet_idx == 0:
                ws = wb.active
                ws.title = _safe_sheet_name(name)
            else:
                ws = wb.create_sheet(title=_safe_sheet_name(name))

            col_labels = [
                str(cd.get("label", cd.get("key", ""))) for cd in col_defs
            ]

            # Write headers
            for col_idx, label in enumerate(col_labels, 1):
                cell = ws.cell(row=1, column=col_idx, value=label)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            # Build per-column number formats
            col_formats: list[str | None] = [
                cd.get("format") if isinstance(cd, dict) else None
                for cd in col_defs
            ]

            # Write data rows
            row_styles: list[dict | None] = opts.get("row_styles", [])
            for row_idx, row_data in enumerate(rows, 2):
                style = (
                    row_styles[row_idx - 2]
                    if row_idx - 2 < len(row_styles)
                    else None
                )
                row_font = _build_row_font(style, default_cell_font)
                row_fill = _build_row_fill(style)

                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(
                        row=row_idx, column=col_idx, value=_coerce_value(value)
                    )
                    cell.font = row_font
                    cell.border = thin_border
                    if row_fill is not None:
                        cell.fill = row_fill
                    # Apply column number format
                    fmt = col_formats[col_idx - 1] if col_idx - 1 < len(col_formats) else None
                    if fmt is not None:
                        cell.number_format = fmt

            # Column widths
            for col_idx, cd in enumerate(col_defs, 1):
                col_letter = get_column_letter(col_idx)
                if isinstance(cd, dict) and "width" in cd:
                    ws.column_dimensions[col_letter].width = cd["width"]
                else:
                    # Auto-width fallback
                    max_len = 0
                    for ws_row in ws.iter_rows(
                        min_col=col_idx, max_col=col_idx
                    ):
                        for cell in ws_row:
                            if cell.value is not None:
                                max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = min(
                        max(max_len + 2, 8), 50
                    )

            # Frozen panes
            frozen_rows = opts.get("frozen_rows")
            if isinstance(frozen_rows, int) and frozen_rows > 0:
                # +1 for header row, +1 because freeze_panes is the first
                # *unfrozen* cell
                ws.freeze_panes = f"A{frozen_rows + 2}"

            # Auto-filter
            if opts.get("auto_filter") and col_defs:
                last_col_letter = get_column_letter(len(col_defs))
                ws.auto_filter.ref = f"A1:{last_col_letter}{len(rows) + 1}"

    else:
        # Legacy single-sheet path
        columns, rows = _normalize_spec(spec)
        title = spec.get("title", "Sheet1")

        ws = wb.active
        ws.title = _safe_sheet_name(title or "Sheet1")

        # Write headers
        if columns:
            for col_idx, label in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=label)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

        # Write data rows
        start_row = 2 if columns else 1
        for row_idx, row_data in enumerate(rows, start_row):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(
                    row=row_idx, column=col_idx, value=_coerce_value(value)
                )
                cell.font = default_cell_font
                cell.border = thin_border

        # Auto-width columns (approximate)
        for col_idx in range(1, (len(columns) or 1) + 1):
            max_len = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(
                max(max_len + 2, 8), 50
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

def export_csv(spec: dict) -> bytes:
    """Generate CSV from a table spec.

    For multi-sheet specs, only the first sheet is exported (CSV does not
    support multiple sheets).

    Args:
        spec: Table spec with ``columns``/``rows`` or ``sheets``.

    Returns:
        UTF-8 encoded CSV content as bytes.
    """
    if _is_multi_sheet(spec):
        # CSV only supports a single sheet — export the first one
        sheets_data = _normalize_multi_sheet(spec)
        if not sheets_data:
            return b""
        _name, col_defs, rows, _opts = sheets_data[0]
        columns = [
            str(cd.get("label", cd.get("key", ""))) for cd in col_defs
        ]
    else:
        columns, rows = _normalize_spec(spec)

    buf = io.StringIO()
    writer = csv.writer(buf)
    if columns:
        writer.writerow(columns)
    for row in rows:
        writer.writerow([_coerce_value(v) for v in row])

    return buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _safe_sheet_name(name: str) -> str:
    """Sanitise a string for use as an Excel sheet name (max 31 chars)."""
    safe = name or "Sheet1"
    for ch in r'\/*?:[]':
        safe = safe.replace(ch, "_")
    return safe[:31]


def _build_row_font(style: dict | None, default: Any) -> Any:
    """Build an openpyxl ``Font`` from a ``_style`` dict, or return *default*."""
    if style is None:
        return default

    from openpyxl.styles import Font

    bold = style.get("bold", False)
    italic = style.get("italic", False)
    font_color = style.get("font_color")
    font_size = style.get("font_size")

    # Only create a custom font if something differs from the default
    if not bold and not italic and font_color is None and font_size is None:
        return default

    color = font_color.lstrip("#") if isinstance(font_color, str) else None
    return Font(
        name="Calibri",
        size=font_size if isinstance(font_size, (int, float)) else 11,
        bold=bool(bold),
        italic=bool(italic),
        color=color,
    )


def _build_row_fill(style: dict | None) -> Any:
    """Build an openpyxl ``PatternFill`` from a ``_style`` dict, or ``None``."""
    if style is None:
        return None

    bg_color = style.get("bg_color")
    if not isinstance(bg_color, str):
        return None

    from openpyxl.styles import PatternFill

    color = bg_color.lstrip("#")
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _coerce_value(value: Any) -> Any:
    """Coerce a cell value to a native Python type for Excel/CSV.

    Attempts numeric conversion for string values that look like numbers.
    """
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    s = str(value).strip()
    if not s:
        return ""
    # Try int
    try:
        return int(s)
    except (ValueError, OverflowError):
        pass
    # Try float
    try:
        return float(s)
    except (ValueError, OverflowError):
        pass
    return s
