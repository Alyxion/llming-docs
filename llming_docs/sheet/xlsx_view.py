"""Read API for the XLSX-backed ``table`` doc type.

Four functions:

* :func:`full_view` — *not* an MCP tool. Generates the full renderable
  snapshot the client ships to the browser — same shape the legacy
  JSON-storage format used (`{sheets: [{name, columns, rows, cells?}]}`)
  so the existing frontend table renderer can consume it without
  vendoring an XLSX JS library. Derived from the workbook on demand;
  not persisted as the canonical storage.

Three more, each backing one MCP tool:

* :func:`overview` — for ``read_document``: cheap "what's this doc?"
  summary (sheet names, dimensions, headers, sample rows).
* :func:`query_table` — for ``query_table``: returns rows in the
  llming-flux ``TableQueryResult`` shape so the two systems are
  interchangeable from the LLM's perspective.
* :func:`query_cells` — for ``query_cells``: sparse cell-level access
  with formatting (XLSX-only, no flux equivalent).

Pagination defaults match flux: ``limit=20``, max ``100``, ``offset``
supported. Filters are case-insensitive substring matches per column —
also identical to flux semantics.
"""
from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string, range_boundaries


# llming-flux uses these defaults; we mirror them so the tools are drop-in
# compatible from the LLM's perspective.
_DEFAULT_LIMIT = 20
_MAX_LIMIT = 100
_SAMPLE_ROWS_IN_OVERVIEW = 5


# ---------------------------------------------------------------------------
# full_view — client render snapshot (not an MCP tool)
# ---------------------------------------------------------------------------


def full_view(wb: Workbook) -> dict:
    """Return the complete renderable snapshot of the workbook.

    Shape::

        {
          "sheets": [{
            "name":    "...",
            "columns": [<header row 1>, ...],   # display labels
            "rows":    [[<row 2>], [<row 3>], ...],
            "cells":   {"B3": {value, font?, fill?, ...}},  # SPARSE — only
                                                            # cells with explicit
                                                            # formatting
            "column_widths": {"B": 18},
            "row_heights":   {3: 22},
            "merges":  ["A1:B2"],
            "freeze_panes": "A2",
          }]
        }

    Generated from the workbook on demand; never persisted. The frontend
    table plugin reads this directly so it doesn't need a JS XLSX
    library — keeps the editor lightweight and easy to embed in other
    hosts (e.g. a Quasar app).
    """
    sheets_out = []
    for ws in wb.worksheets:
        sheets_out.append(_sheet_full_view(ws))
    return {"sheets": sheets_out}


def _sheet_full_view(ws) -> dict:
    """Per-sheet snapshot — same shape ``full_view`` yields, scoped to one sheet."""
    out: dict = {"name": ws.title}

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    # Header row.
    columns: list[Any] = [
        ws.cell(row=1, column=c).value for c in range(1, max_col + 1)
    ]
    while columns and columns[-1] in (None, ""):
        columns.pop()
    out["columns"] = columns

    eff_cols = max(len(columns), 1)

    # Data rows.
    rows: list[list[Any]] = []
    for r in range(2, max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, eff_cols + 1)]
        if all(v in (None, "") for v in row):
            continue
        rows.append(row)
    out["rows"] = rows

    # Sparse per-cell formatting / value overrides.
    cells_out: dict[str, dict] = {}
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            entry = _cell_to_dict(cell, with_formatting=True)
            if entry and (entry.get("font") or entry.get("fill")
                          or entry.get("alignment") or entry.get("border")
                          or entry.get("number_format")):
                # Only emit when there's actual formatting — values alone
                # are already in ``rows`` / ``columns``, no need to duplicate.
                cells_out[cell.coordinate] = entry
    if cells_out:
        out["cells"] = cells_out

    # Column widths / row heights — only explicit (customWidth/customHeight).
    col_widths: dict[str, float] = {}
    for letter, dim in ws.column_dimensions.items():
        if dim.customWidth and dim.width is not None:
            col_widths[letter] = float(dim.width)
    if col_widths:
        out["column_widths"] = col_widths

    row_heights: dict[int, float] = {}
    for r_idx, dim in ws.row_dimensions.items():
        if dim.customHeight and dim.height is not None:
            row_heights[int(r_idx)] = float(dim.height)
    if row_heights:
        out["row_heights"] = row_heights

    merges = sorted(str(rng) for rng in ws.merged_cells.ranges)
    if merges:
        out["merges"] = merges

    if ws.freeze_panes:
        out["freeze_panes"] = ws.freeze_panes

    return out


# ---------------------------------------------------------------------------
# overview — for read_document
# ---------------------------------------------------------------------------


def overview(wb: Workbook, name: str = "") -> dict:
    """Return a small "describe this workbook" dict.

    Used by ``read_document(document_id)``. The shape is type-discriminated
    at the tool layer; this function produces just the table-specific body.
    """
    sheets_out = []
    for ws in wb.worksheets:
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        headers = [
            ws.cell(row=1, column=c).value
            for c in range(1, max_col + 1)
        ]
        # Trim trailing all-None header cells so the JSON stays tight.
        while headers and headers[-1] in (None, ""):
            headers.pop()
        eff_cols = max(len(headers), 1)

        # Up to N sample rows below the header to give the LLM a feel
        # for the data without exploding token cost on huge sheets.
        sample = []
        for r in range(2, min(max_row, 1 + _SAMPLE_ROWS_IN_OVERVIEW) + 1):
            row = {
                str(headers[c - 1]) if c - 1 < len(headers) and headers[c - 1] is not None
                else f"col_{c}":
                ws.cell(row=r, column=c).value
                for c in range(1, eff_cols + 1)
            }
            sample.append(row)

        sheets_out.append({
            "name": ws.title,
            "max_row": max_row,
            "max_col": max_col,
            "headers": headers,
            "sample_rows": sample,
        })
    return {
        "type": "table",
        "name": name,
        "sheets": sheets_out,
    }


# ---------------------------------------------------------------------------
# query_table — flux-compatible row query
# ---------------------------------------------------------------------------


def query_table(
    wb: Workbook,
    sheet: Any = None,
    range_addr: str | None = None,
    columns: list[str] | None = None,
    filters: dict[str, Any] | None = None,
    limit: int | None = None,
    offset: int | None = None,
) -> dict:
    """Return rows from a sheet in the llming-flux ``TableQueryResult``
    shape::

        {
          "table_name": "...",
          "headers":    [str, ...],
          "rows":       [{col: value}, ...],
          "total_rows": int,   # rows after filters, before limit/offset
          "returned_rows": int,
        }

    ``sheet`` is name-or-index; defaults to the first. ``range_addr`` is
    A1-style and limits both the column projection and row range. The
    header row is **always** XLSX row 1 within the projected columns —
    matching how every spreadsheet user reads a table.
    """
    ws = _resolve_sheet(wb, sheet)

    # Resolve the rectangular slab we'll read.
    if range_addr:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(range_addr)
        except (ValueError, TypeError):
            raise ValueError(f"'{range_addr}' is not a valid range (e.g. 'A1:E50')")
    else:
        min_col, max_col = 1, max(ws.max_column or 1, 1)
        min_row, max_row = 1, max(ws.max_row or 1, 1)

    # Header row is the first row of the slab.
    header_row = min_row
    headers_raw = [
        ws.cell(row=header_row, column=c).value
        for c in range(min_col, max_col + 1)
    ]
    # Display labels — fall back to "col_<n>" for blank header cells so
    # row dicts always have a usable key.
    headers = [
        (str(h) if h not in (None, "") else f"col_{c}")
        for c, h in zip(range(min_col, max_col + 1), headers_raw)
    ]

    # Optional column projection: keep only the requested ones, in the
    # requested order.
    keep_cols = list(range(min_col, max_col + 1))
    keep_headers = headers
    if columns:
        wanted = [str(c) for c in columns]
        new_keep = []
        new_headers = []
        for w in wanted:
            for hdr, col in zip(headers, keep_cols):
                if hdr == w:
                    new_keep.append(col)
                    new_headers.append(hdr)
                    break
        keep_cols = new_keep
        keep_headers = new_headers

    # Build all data rows first, then apply filters, then paginate.
    all_rows: list[dict[str, Any]] = []
    for r in range(header_row + 1, max_row + 1):
        row = {hdr: ws.cell(row=r, column=col).value
               for hdr, col in zip(keep_headers, keep_cols)}
        if all(v in (None, "") for v in row.values()):
            continue  # skip fully-empty rows
        all_rows.append(row)

    # Filters: case-insensitive substring contains, identical to flux.
    if filters:
        def matches(row):
            for key, needle in filters.items():
                if key not in row:
                    return False
                hay = "" if row[key] is None else str(row[key])
                if str(needle).lower() not in hay.lower():
                    return False
            return True
        all_rows = [r for r in all_rows if matches(r)]

    total_rows = len(all_rows)
    off = max(int(offset or 0), 0)
    lim = max(min(int(limit or _DEFAULT_LIMIT), _MAX_LIMIT), 1)
    page = all_rows[off:off + lim]

    return {
        "table_name": ws.title,
        "headers": keep_headers,
        "rows": page,
        "total_rows": total_rows,
        "returned_rows": len(page),
    }


# ---------------------------------------------------------------------------
# query_cells — sparse cell-level + formatting
# ---------------------------------------------------------------------------


def query_cells(
    wb: Workbook,
    sheet: Any,
    range_addr: str,
    with_formatting: bool = True,
) -> dict:
    """Return cells in a range, sparse, with optional formatting.

    Useful when the LLM is inspecting a specific area for styles, number
    formats, or comparing cells. For bulk data reads, use
    :func:`query_table` — its row-shaped output is cheaper.

    Empty cells (no value, no formatting) are omitted to keep the
    payload small.
    """
    ws = _resolve_sheet(wb, sheet)
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_addr)
    except (ValueError, TypeError):
        raise ValueError(f"'{range_addr}' is not a valid range (e.g. 'A1:E10')")

    cells_out: dict[str, dict] = {}
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            entry = _cell_to_dict(cell, with_formatting=with_formatting)
            if entry:
                cells_out[cell.coordinate] = entry

    return {
        "sheet_name": ws.title,
        "range": range_addr,
        "cells": cells_out,
    }


# ---------------------------------------------------------------------------
# Cell → dict (sparse, openpyxl-shaped)
# ---------------------------------------------------------------------------


def _cell_to_dict(cell: Cell, with_formatting: bool) -> dict:
    """Emit only the fields that are non-default. An empty result means
    the cell is fully blank and the caller can skip emitting it."""
    out: dict = {}
    if cell.value is not None and cell.value != "":
        out["value"] = cell.value

    if not with_formatting:
        return out

    if cell.number_format and cell.number_format != "General":
        out["number_format"] = cell.number_format

    font = _font_to_dict(cell.font)
    if font:
        out["font"] = font

    fill = _fill_to_dict(cell.fill)
    if fill:
        out["fill"] = fill

    align = _alignment_to_dict(cell.alignment)
    if align:
        out["alignment"] = align

    border = _border_to_dict(cell.border)
    if border:
        out["border"] = border

    return out


def _font_to_dict(font) -> dict:
    if font is None:
        return {}
    out: dict = {}
    if font.bold:
        out["bold"] = True
    if font.italic:
        out["italic"] = True
    if font.underline and font.underline != "none":
        out["underline"] = font.underline
    if font.strike:
        out["strike"] = True
    if font.size and float(font.size) != 11.0:
        out["size"] = float(font.size)
    if font.name and font.name != "Calibri":
        out["name"] = font.name
    color_hex = _color_hex(font.color)
    if color_hex and color_hex.upper() not in ("000000", "FF000000"):
        out["color"] = color_hex
    return out


def _fill_to_dict(fill) -> dict:
    if fill is None or not fill.fill_type or fill.fill_type == "none":
        return {}
    out: dict = {"fill_type": fill.fill_type}
    start = _color_hex(fill.start_color)
    if start:
        out["start_color"] = start
    end = _color_hex(fill.end_color)
    if end and end != start:
        out["end_color"] = end
    return out


def _alignment_to_dict(align) -> dict:
    if align is None:
        return {}
    out: dict = {}
    if align.horizontal and align.horizontal != "general":
        out["horizontal"] = align.horizontal
    if align.vertical and align.vertical != "bottom":
        out["vertical"] = align.vertical
    if align.wrap_text:
        out["wrap_text"] = True
    if align.shrink_to_fit:
        out["shrink_to_fit"] = True
    if align.indent:
        out["indent"] = align.indent
    if align.text_rotation:
        out["text_rotation"] = align.text_rotation
    return out


def _border_to_dict(border) -> dict:
    if border is None:
        return {}
    out: dict = {}
    for name in ("left", "right", "top", "bottom", "diagonal"):
        side = getattr(border, name, None)
        if side is None or not getattr(side, "style", None):
            continue
        side_out = {"style": side.style}
        color = _color_hex(side.color)
        if color:
            side_out["color"] = color
        out[name] = side_out
    return out


def _color_hex(color) -> str | None:
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str) and len(rgb) >= 6:
        return rgb[-6:].upper()
    return None


# ---------------------------------------------------------------------------
# Sheet resolution (shared)
# ---------------------------------------------------------------------------


def _resolve_sheet(wb: Workbook, sheet: Any):
    if sheet is None:
        return wb.worksheets[0] if wb.worksheets else wb.active
    if isinstance(sheet, int):
        if sheet < 0 or sheet >= len(wb.worksheets):
            raise ValueError(
                f"sheet index {sheet} out of range "
                f"(workbook has {len(wb.worksheets)} sheet(s))"
            )
        return wb.worksheets[sheet]
    if isinstance(sheet, str):
        if sheet.isdigit():
            return _resolve_sheet(wb, int(sheet))
        if sheet in wb.sheetnames:
            return wb[sheet]
        raise ValueError(
            f"no sheet named '{sheet}'; existing: {wb.sheetnames}"
        )
    raise ValueError(f"sheet must be a name (str) or 0-based index (int)")
