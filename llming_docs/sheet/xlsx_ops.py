"""Apply LLM-emitted JSON ops to an openpyxl ``Workbook``.

The op vocabulary is documented in ``docs/api-design.md``; the short
version is openpyxl-native A1 paths:

    set    sheets/<i>/cells/<addr>/value
    set    sheets/<i>/cells/<addr>/font          {bold, italic, color, size, name}
    set    sheets/<i>/cells/<addr>/fill          {fill_type, start_color, end_color}
    set    sheets/<i>/cells/<addr>/alignment     {horizontal, vertical, wrap_text}
    set    sheets/<i>/cells/<addr>/border        {top, right, bottom, left, diagonal}
    set    sheets/<i>/cells/<addr>/number_format "<format string>"
    add    sheets/<i>/rows/-                     value: [v1, v2, ...]
    add    sheets/<i>/rows                       value: [...] position: <r>
    remove sheets/<i>/rows/<r>
    add    sheets/<i>/columns/-                  value: "<header>"
    remove sheets/<i>/columns/<addr_or_index>
    set    sheets/<i>/columns/<col>/width        value: <number>
    set    sheets/<i>/rows/<r>/height            value: <number>
    add    sheets/-                              value: "<sheet name>"
    remove sheets/<i>
    set    sheets/<i>/name                       value: "<new name>"
    add    sheets/<i>/merges/-                   value: "A1:B2"
    set    sheets/<i>/freeze_panes               value: "A2"
    bulk_set sheets/<i>/range/<top_left>         values: [[...], [...]]

Every handler raises ``XlsxOpError`` on bad input — the caller surfaces
the message back to the LLM so it can self-correct.
"""
from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string


class XlsxOpError(ValueError):
    """Raised by ``apply_op`` when an op cannot be applied. The message
    is plain enough for the LLM to read and try a different path / value
    on the retry."""


# ---------------------------------------------------------------------------
# Public entry points
# ---------------------------------------------------------------------------


def apply_op(wb: Workbook, op: dict) -> None:
    """Apply a single op (mutates ``wb`` in place)."""
    if not isinstance(op, dict):
        raise XlsxOpError(f"op must be a dict, got {type(op).__name__}")
    op_type = op.get("op")
    path = op.get("path", "")
    if not op_type:
        raise XlsxOpError("op is missing the 'op' field")
    if not isinstance(path, str):
        raise XlsxOpError("op 'path' must be a string")

    segments = [s for s in path.split("/") if s != ""]
    # Trailing slash on inputs like "sheets/-/" is tolerated; an empty
    # path means "operate on the workbook root" — only ``add sheets/-``
    # actually uses that today.

    handler = _OP_HANDLERS.get(op_type)
    if handler is None:
        raise XlsxOpError(f"unknown op type '{op_type}'")
    handler(wb, segments, op)


def apply_operations(wb: Workbook, ops: list) -> None:
    """Apply a batch — fail-fast: the first failure aborts the rest. The
    caller is expected to deep-copy the workbook before calling, so a
    half-applied batch can be discarded cleanly."""
    if not isinstance(ops, list):
        raise XlsxOpError("operations must be a list")
    for i, op in enumerate(ops):
        try:
            apply_op(wb, op)
        except XlsxOpError as exc:
            raise XlsxOpError(f"operation {i}: {exc}") from exc


# ---------------------------------------------------------------------------
# Path resolution helpers
# ---------------------------------------------------------------------------


def _resolve_sheet(wb: Workbook, segment: str):
    """``sheets/<i>`` — accept 0-based index OR sheet name."""
    if segment.isdigit():
        idx = int(segment)
        if idx < 0 or idx >= len(wb.worksheets):
            raise XlsxOpError(
                f"sheet index {idx} out of range (workbook has "
                f"{len(wb.worksheets)} sheet(s))"
            )
        return wb.worksheets[idx]
    if segment in wb.sheetnames:
        return wb[segment]
    raise XlsxOpError(
        f"no sheet named or indexed '{segment}'; "
        f"existing: {wb.sheetnames}"
    )


def _resolve_sheet_index(wb: Workbook, segment: str) -> int:
    """Like ``_resolve_sheet`` but returns the 0-based index. Needed by
    ``remove sheets/<i>`` since openpyxl removes by sheet object."""
    if segment.isdigit():
        idx = int(segment)
        if idx < 0 or idx >= len(wb.worksheets):
            raise XlsxOpError(f"sheet index {idx} out of range")
        return idx
    if segment in wb.sheetnames:
        return wb.sheetnames.index(segment)
    raise XlsxOpError(f"no sheet named or indexed '{segment}'")


def _column_index(segment: str) -> int:
    """Accept ``B`` (letter) or ``2`` (1-based index) as a column ref."""
    if segment.isdigit():
        idx = int(segment)
        if idx < 1:
            raise XlsxOpError(f"column index must be >= 1, got {idx}")
        return idx
    try:
        return column_index_from_string(segment.upper())
    except (ValueError, KeyError):
        raise XlsxOpError(f"'{segment}' is not a valid column letter or index")


def _row_index(segment: str) -> int:
    if not segment.isdigit():
        raise XlsxOpError(f"row index must be a number, got '{segment}'")
    idx = int(segment)
    if idx < 1:
        raise XlsxOpError(f"row index must be >= 1, got {idx}")
    return idx


# ---------------------------------------------------------------------------
# Cell-level ops
# ---------------------------------------------------------------------------


_FONT_KEYS = ("name", "size", "bold", "italic", "underline", "strike", "color")
_ALIGN_KEYS = ("horizontal", "vertical", "wrap_text", "shrink_to_fit",
               "indent", "text_rotation")
_FILL_KEYS = ("fill_type", "start_color", "end_color")
_SIDE_KEYS = ("style", "color")
_BORDER_SIDES = ("left", "right", "top", "bottom", "diagonal")


def _cell_at(wb: Workbook, segs: list[str]) -> Cell:
    """Resolve ``sheets/<i>/cells/<addr>`` to the underlying cell.

    Caller passes ``segs`` already split; we expect at minimum
    ``["sheets", "<i>", "cells", "<addr>"]`` and accept any extra trailing
    segments (the caller decides what to do with them — usually a sub-
    field like ``value`` / ``font`` / etc.).
    """
    if len(segs) < 4 or segs[0] != "sheets" or segs[2] != "cells":
        raise XlsxOpError(
            f"expected 'sheets/<i>/cells/<addr>/...', got '{'/'.join(segs)}'"
        )
    ws = _resolve_sheet(wb, segs[1])
    addr = segs[3].upper()
    # Be defensive — openpyxl raises a CellCoordinatesException (subclass
    # of Exception, not ValueError) for some malformed inputs, and silently
    # accepts letters-without-digits as a "column only" coordinate.
    try:
        col_letter, row_num = coordinate_from_string(addr)
    except Exception:
        raise XlsxOpError(f"'{addr}' is not a valid cell address (e.g. 'B3')")
    if not isinstance(row_num, int) or row_num < 1 or not col_letter:
        raise XlsxOpError(f"'{addr}' is not a valid cell address (e.g. 'B3')")
    return ws.cell(row=row_num, column=column_index_from_string(col_letter))


def _set_cell_value(cell: Cell, value: Any) -> None:
    cell.value = _coerce_value(value)


def _set_cell_font(cell: Cell, value: Any) -> None:
    if not isinstance(value, dict):
        raise XlsxOpError("'font' value must be an object")
    kwargs = {k: value[k] for k in _FONT_KEYS if k in value}
    if not kwargs:
        # Reset to default (allows "remove styling" via empty object).
        cell.font = Font()
        return
    cell.font = Font(**kwargs)


def _set_cell_fill(cell: Cell, value: Any) -> None:
    if not isinstance(value, dict):
        raise XlsxOpError("'fill' value must be an object")
    kwargs = {k: value[k] for k in _FILL_KEYS if k in value}
    if "start_color" in kwargs and "fill_type" not in kwargs:
        # Common case: caller wants a solid background, doesn't know about
        # ``fill_type``. Default to solid.
        kwargs["fill_type"] = "solid"
    if not kwargs:
        cell.fill = PatternFill(fill_type=None)
        return
    cell.fill = PatternFill(**kwargs)


def _set_cell_alignment(cell: Cell, value: Any) -> None:
    if not isinstance(value, dict):
        raise XlsxOpError("'alignment' value must be an object")
    kwargs = {k: value[k] for k in _ALIGN_KEYS if k in value}
    cell.alignment = Alignment(**kwargs) if kwargs else Alignment()


def _set_cell_border(cell: Cell, value: Any) -> None:
    if not isinstance(value, dict):
        raise XlsxOpError("'border' value must be an object")
    sides: dict[str, Side] = {}
    for side_name in _BORDER_SIDES:
        side_def = value.get(side_name)
        if side_def is None:
            continue
        if not isinstance(side_def, dict):
            raise XlsxOpError(f"border '{side_name}' must be an object")
        side_kwargs = {k: side_def[k] for k in _SIDE_KEYS if k in side_def}
        sides[side_name] = Side(**side_kwargs)
    cell.border = Border(**sides) if sides else Border()


def _set_cell_number_format(cell: Cell, value: Any) -> None:
    if not isinstance(value, str):
        raise XlsxOpError("'number_format' must be a string")
    cell.number_format = value


_CELL_FIELD_HANDLERS = {
    "value": _set_cell_value,
    "font": _set_cell_font,
    "fill": _set_cell_fill,
    "alignment": _set_cell_alignment,
    "border": _set_cell_border,
    "number_format": _set_cell_number_format,
}


def _set_cell_bundle(cell: Cell, value: Any) -> None:
    """``set sheets/<i>/cells/<addr>`` (no field) — value is a dict of
    cell properties (any subset of the field handlers). Useful when the
    LLM wants to set value + format in one op."""
    if not isinstance(value, dict):
        raise XlsxOpError(
            "setting a cell directly requires an object with any of "
            f"{list(_CELL_FIELD_HANDLERS)}"
        )
    for field, handler in _CELL_FIELD_HANDLERS.items():
        if field in value:
            handler(cell, value[field])


# ---------------------------------------------------------------------------
# `set` op
# ---------------------------------------------------------------------------


def _op_set(wb: Workbook, segs: list[str], op: dict) -> None:
    if "value" not in op:
        raise XlsxOpError("'set' operation requires a 'value' field")
    value = op["value"]

    # Workbook-level paths first ---------------------------------------
    if not segs:
        raise XlsxOpError("'set' requires a path")

    if segs[0] != "sheets":
        raise XlsxOpError(f"only 'sheets/...' paths are supported, got '{segs[0]}'")

    if len(segs) == 1:
        raise XlsxOpError("path 'sheets' alone is not settable; use 'sheets/<i>/...'")

    # sheets/<i>/...
    sheet_seg = segs[1]
    if len(segs) == 2:
        # Setting an entire sheet — out of scope; use ops to populate.
        raise XlsxOpError(
            "cannot set a whole sheet; use individual ops like "
            "sheets/<i>/cells/<addr>/value"
        )
    field = segs[2]

    if field == "name":
        ws = _resolve_sheet(wb, sheet_seg)
        if not isinstance(value, str) or not value.strip():
            raise XlsxOpError("sheet name must be a non-empty string")
        ws.title = value
        return

    if field == "freeze_panes":
        ws = _resolve_sheet(wb, sheet_seg)
        if not isinstance(value, str):
            raise XlsxOpError("freeze_panes must be a string like 'A2'")
        ws.freeze_panes = value
        return

    if field == "cells":
        if len(segs) < 4:
            raise XlsxOpError("'cells' path needs an address: cells/<addr>")
        cell = _cell_at(wb, segs)
        # sheets/<i>/cells/<addr>            → bundle set
        # sheets/<i>/cells/<addr>/<subfield> → field set
        if len(segs) == 4:
            _set_cell_bundle(cell, value)
            return
        subfield = segs[4]
        handler = _CELL_FIELD_HANDLERS.get(subfield)
        if handler is None:
            raise XlsxOpError(
                f"unknown cell field '{subfield}'; "
                f"valid: {sorted(_CELL_FIELD_HANDLERS)}"
            )
        handler(cell, value)
        return

    if field == "columns":
        # sheets/<i>/columns/<col>/<attr>
        if len(segs) < 5:
            raise XlsxOpError("column path needs <col>/<attr>")
        ws = _resolve_sheet(wb, sheet_seg)
        col_letter = get_column_letter(_column_index(segs[3]))
        attr = segs[4]
        if attr == "width":
            try:
                ws.column_dimensions[col_letter].width = float(value)
            except (TypeError, ValueError):
                raise XlsxOpError(f"column width must be numeric, got {value!r}")
            return
        raise XlsxOpError(f"unknown column attribute '{attr}'")

    if field == "rows":
        # sheets/<i>/rows/<r>/<attr>
        if len(segs) < 5:
            raise XlsxOpError("row path needs <r>/<attr>")
        ws = _resolve_sheet(wb, sheet_seg)
        row_idx = _row_index(segs[3])
        attr = segs[4]
        if attr == "height":
            try:
                ws.row_dimensions[row_idx].height = float(value)
            except (TypeError, ValueError):
                raise XlsxOpError(f"row height must be numeric, got {value!r}")
            return
        raise XlsxOpError(f"unknown row attribute '{attr}'")

    raise XlsxOpError(f"unknown 'set' target 'sheets/<i>/{field}'")


# ---------------------------------------------------------------------------
# `add` op
# ---------------------------------------------------------------------------


def _op_add(wb: Workbook, segs: list[str], op: dict) -> None:
    if "value" not in op:
        raise XlsxOpError("'add' operation requires a 'value' field")
    value = op["value"]
    position = op.get("position")

    if not segs or segs[0] != "sheets":
        raise XlsxOpError("'add' supports 'sheets/...' paths only")

    # sheets/-                  — append a new sheet
    # sheets/<i>/rows/-         — append row
    # sheets/<i>/rows           with position — insert row at position
    # sheets/<i>/columns/-      — append column header
    # sheets/<i>/merges/-       — add a merge range

    if len(segs) == 2 and segs[1] == "-":
        # Append sheet. value: name (str) OR null/dict (use default).
        name = value if isinstance(value, str) and value.strip() else None
        if name and name in wb.sheetnames:
            raise XlsxOpError(f"sheet '{name}' already exists")
        ws = wb.create_sheet(title=name)
        return

    sheet_seg = segs[1]
    field = segs[2] if len(segs) > 2 else None

    if field == "rows":
        ws = _resolve_sheet(wb, sheet_seg)
        if len(segs) == 3:
            # sheets/<i>/rows with explicit position
            if position is None:
                raise XlsxOpError("'add sheets/<i>/rows' requires 'position'")
            row_idx = int(position)
            if row_idx < 1:
                raise XlsxOpError("row position must be >= 1")
            ws.insert_rows(row_idx)
            _write_row_values(ws, row_idx, value)
            return
        if len(segs) == 4 and segs[3] == "-":
            # Append row at max_row + 1
            row_idx = (ws.max_row or 0) + 1
            _write_row_values(ws, row_idx, value)
            return

    if field == "columns":
        ws = _resolve_sheet(wb, sheet_seg)
        if len(segs) == 4 and segs[3] == "-":
            # Append column — set header in row 1 at max_column + 1
            col_idx = (ws.max_column or 0) + 1
            ws.cell(row=1, column=col_idx, value=_coerce_value(value))
            return
        # Insert at index
        if len(segs) == 3 and position is not None:
            col_idx = _column_index(str(position))
            ws.insert_cols(col_idx)
            ws.cell(row=1, column=col_idx, value=_coerce_value(value))
            return

    if field == "merges":
        ws = _resolve_sheet(wb, sheet_seg)
        if len(segs) == 4 and segs[3] == "-":
            if not isinstance(value, str):
                raise XlsxOpError("merge value must be a range string like 'A1:B2'")
            try:
                ws.merge_cells(value)
            except (ValueError, TypeError) as exc:
                raise XlsxOpError(f"invalid merge range '{value}': {exc}")
            return

    raise XlsxOpError(f"unknown 'add' target '{'/'.join(segs)}'")


def _write_row_values(ws, row_idx: int, value: Any) -> None:
    """Write a list-of-values into row ``row_idx`` starting at column 1."""
    if not isinstance(value, list):
        raise XlsxOpError("row value must be a list of cell values")
    for c, v in enumerate(value, start=1):
        ws.cell(row=row_idx, column=c, value=_coerce_value(v))


# ---------------------------------------------------------------------------
# `remove` op
# ---------------------------------------------------------------------------


def _op_remove(wb: Workbook, segs: list[str], op: dict) -> None:
    if not segs or segs[0] != "sheets":
        raise XlsxOpError("'remove' supports 'sheets/...' paths only")

    if len(segs) == 2:
        # remove sheets/<i>
        if len(wb.worksheets) <= 1:
            raise XlsxOpError("cannot remove the last sheet")
        idx = _resolve_sheet_index(wb, segs[1])
        wb.remove(wb.worksheets[idx])
        return

    sheet_seg = segs[1]
    field = segs[2]

    if field == "rows" and len(segs) >= 4:
        ws = _resolve_sheet(wb, sheet_seg)
        row_idx = _row_index(segs[3])
        ws.delete_rows(row_idx)
        return

    if field == "columns" and len(segs) >= 4:
        ws = _resolve_sheet(wb, sheet_seg)
        col_idx = _column_index(segs[3])
        ws.delete_cols(col_idx)
        return

    if field == "cells" and len(segs) >= 4:
        # Removing a cell == clearing its value + style.
        cell = _cell_at(wb, segs[:4])
        cell.value = None
        cell.font = Font()
        cell.fill = PatternFill(fill_type=None)
        cell.alignment = Alignment()
        cell.border = Border()
        cell.number_format = "General"
        return

    if field == "merges" and len(segs) >= 4:
        ws = _resolve_sheet(wb, sheet_seg)
        try:
            ws.unmerge_cells(segs[3])
        except (ValueError, TypeError) as exc:
            raise XlsxOpError(f"cannot unmerge '{segs[3]}': {exc}")
        return

    raise XlsxOpError(f"unknown 'remove' target '{'/'.join(segs)}'")


# ---------------------------------------------------------------------------
# `bulk_set` op — sugar for "fill a rectangular block starting at <addr>"
# ---------------------------------------------------------------------------


def _op_bulk_set(wb: Workbook, segs: list[str], op: dict) -> None:
    """``bulk_set sheets/<i>/range/<top_left>`` with ``values: [[...]]``.

    Drops a 2D array of values into the sheet starting at ``<top_left>``.
    Equivalent to a sequence of per-cell ``set`` ops; the LLM uses this
    when drafting larger tables to save tokens.
    """
    if "values" not in op:
        raise XlsxOpError("'bulk_set' requires a 'values' field (2D array)")
    values = op["values"]
    if not isinstance(values, list) or not all(isinstance(r, list) for r in values):
        raise XlsxOpError("'bulk_set' values must be a 2D array")

    if len(segs) < 4 or segs[0] != "sheets" or segs[2] != "range":
        raise XlsxOpError(
            "expected 'sheets/<i>/range/<top_left>', got "
            f"'{'/'.join(segs)}'"
        )
    ws = _resolve_sheet(wb, segs[1])
    top_left = segs[3].upper()
    try:
        col_letter, base_row = coordinate_from_string(top_left)
    except Exception:
        raise XlsxOpError(f"'{top_left}' is not a valid cell address")
    base_col = column_index_from_string(col_letter)

    # Use distinct names for the row counter and the row payload so the
    # ``base_row`` integer doesn't get shadowed by the inner row list.
    for r_off, row_values in enumerate(values):
        for c_off, val in enumerate(row_values):
            ws.cell(
                row=base_row + r_off,
                column=base_col + c_off,
                value=_coerce_value(val),
            )


# ---------------------------------------------------------------------------
# Dispatcher table
# ---------------------------------------------------------------------------


_OP_HANDLERS = {
    "set": _op_set,
    "add": _op_add,
    "remove": _op_remove,
    "bulk_set": _op_bulk_set,
}


# ---------------------------------------------------------------------------
# Value coercion
# ---------------------------------------------------------------------------


def _coerce_value(v: Any) -> Any:
    """Pass-through for everything openpyxl handles directly. Lists /
    dicts get JSON-stringified — defensive against the LLM dropping a
    structured value into a cell."""
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, (list, dict)):
        import json as _json
        return _json.dumps(v, ensure_ascii=False)
    return v
