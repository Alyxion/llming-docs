"""openpyxl ``Workbook`` lifecycle — create empty, serialize, load.

The ``table`` document type holds a ``Workbook`` in memory and persists
it as XLSX bytes (base64 in the doc store). This module is the only
place that crosses the bytes ⇄ Workbook boundary; everything else
(`xlsx_ops`, `xlsx_view`) operates on a live Workbook.

Formulas round-trip as text. ``data_only=False`` on load preserves the
formula expressions; we don't compute them server-side.
"""
from __future__ import annotations

import base64
import io

from openpyxl import Workbook, load_workbook


# ---------------------------------------------------------------------------
# Empty / new
# ---------------------------------------------------------------------------


def new_empty_workbook(sheet_name: str = "Sheet 1") -> Workbook:
    """Return a fresh ``Workbook`` with a single named sheet, no data.

    Used by ``create_document(type=table)`` before any ops are applied. The
    LLM is expected to populate it via the ``operations`` payload.
    """
    wb = Workbook()
    # ``Workbook()`` always creates a sheet titled "Sheet". Rename it
    # so the LLM and UI see something more spreadsheet-shaped from the
    # very first read.
    default = wb.active
    default.title = sheet_name
    return wb


# ---------------------------------------------------------------------------
# Bytes <-> Workbook
# ---------------------------------------------------------------------------


def workbook_to_bytes(wb: Workbook) -> bytes:
    """Serialize ``wb`` to in-memory XLSX bytes.

    Used both for persisting the doc and for the user-initiated download
    flow. openpyxl produces a fully-valid .xlsx archive — the same bytes
    Excel / LibreOffice / Numbers will open natively.
    """
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def workbook_from_bytes(data: bytes) -> Workbook:
    """Inverse of ``workbook_to_bytes``.

    ``data_only=False`` so formula cells round-trip as their expression
    (e.g. ``=SUM(A1:A10)``) rather than the cached computed value. This
    matches the design decision to leave formulas unrendered for v1 —
    we'd rather keep the formula and show it as text than silently lose
    it on save.
    """
    return load_workbook(filename=io.BytesIO(data), data_only=False)


# ---------------------------------------------------------------------------
# Base64 <-> Workbook (JSON-friendly transport)
# ---------------------------------------------------------------------------


def workbook_to_b64(wb: Workbook) -> str:
    """Base64-encoded XLSX bytes — the canonical persisted form for the
    ``table`` doc type. Stored under ``Document.data["xlsx_b64"]``."""
    return base64.b64encode(workbook_to_bytes(wb)).decode("ascii")


def workbook_from_b64(b64: str) -> Workbook:
    """Inverse of ``workbook_to_b64``."""
    return workbook_from_bytes(base64.b64decode(b64))
