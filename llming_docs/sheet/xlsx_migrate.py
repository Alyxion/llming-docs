"""One-shot conversion: legacy ``{sheets: [...]}`` JSON → openpyxl ``Workbook``.

This is **not** a runtime bridge. It runs exactly once per legacy doc, at
the document store's load boundary. After the doc is saved back as XLSX
bytes, the JSON shape is gone and never reconstituted.

The function is deliberately tolerant — historical docs span multiple
schema variants (flat ``{columns, rows}``, multi-sheet ``{sheets: [...]}``,
column entries as strings or as ``{name|label|id}`` dicts). We accept
all of them and produce a clean Workbook with row 1 = headers, rows 2..N
= data, exactly like a freshly-built XLSX.

Per-cell styles, column widths, merges, freeze panes from the legacy
JSON are preserved if present; everything else uses openpyxl defaults.
"""
from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ---------------------------------------------------------------------------
# Detection + conversion
# ---------------------------------------------------------------------------


def is_legacy_json_table(data: Any) -> bool:
    """Recognize the old ``{sheets}`` / ``{columns, rows}`` payloads.

    The new shape is ``{xlsx_b64: "..."}``; anything else looking like a
    table dict triggers the migration path."""
    if not isinstance(data, dict):
        return False
    if "xlsx_b64" in data:
        return False
    return ("sheets" in data) or ("columns" in data) or ("rows" in data)


def migrate_legacy_json_to_workbook(data: Any) -> Workbook:
    """Build a Workbook from a legacy table doc payload. Idempotent —
    callers can call ``is_legacy_json_table`` first to skip the work
    when ``data`` already holds an ``xlsx_b64`` field."""
    wb = Workbook()
    # openpyxl's default sheet — drop it; we rebuild from the legacy data.
    wb.remove(wb.active)

    sheets = _collect_sheets(data)
    if not sheets:
        wb.create_sheet(title="Sheet 1")
        return wb

    for sheet_def in sheets:
        ws = wb.create_sheet(
            title=sheet_def.get("name") or f"Sheet {len(wb.sheetnames) + 1}"
        )
        _populate_from_legacy(ws, sheet_def)
    return wb


# ---------------------------------------------------------------------------
# Schema flattening
# ---------------------------------------------------------------------------


def _collect_sheets(data: dict) -> list[dict]:
    """Coerce the two accepted legacy shapes into a uniform sheet list."""
    if isinstance(data.get("sheets"), list):
        return [s for s in data["sheets"] if isinstance(s, dict)]
    keys = ("columns", "rows", "cells", "column_widths", "row_heights",
            "merges", "freeze_panes")
    flat = {k: data[k] for k in keys if k in data}
    if not flat:
        return []
    flat["name"] = data.get("name") or "Sheet 1"
    return [flat]


# ---------------------------------------------------------------------------
# Per-sheet population
# ---------------------------------------------------------------------------


def _populate_from_legacy(ws, sheet_def: dict) -> None:
    columns = sheet_def.get("columns") or []
    rows = sheet_def.get("rows") or []

    # Header row.
    for c, col in enumerate(columns, start=1):
        ws.cell(row=1, column=c, value=_col_label(col))

    # Data rows under the header.
    for r, row in enumerate(rows):
        excel_row = r + 2
        if isinstance(row, list):
            for c, val in enumerate(row, start=1):
                ws.cell(row=excel_row, column=c, value=_coerce_value(val))
        elif isinstance(row, dict):
            keys = [_col_label(c) for c in columns] if columns else list(row.keys())
            for c, key in enumerate(keys, start=1):
                if key in row:
                    ws.cell(row=excel_row, column=c, value=_coerce_value(row[key]))

    # Sparse per-cell overrides.
    cells = sheet_def.get("cells") or {}
    if isinstance(cells, dict):
        for addr, props in cells.items():
            if not isinstance(props, dict):
                continue
            try:
                cell = ws[addr]
            except (ValueError, KeyError):
                continue
            _apply_legacy_cell_props(cell, props)

    # Column widths.
    cw = sheet_def.get("column_widths") or {}
    if isinstance(cw, dict):
        for letter, width in cw.items():
            try:
                ws.column_dimensions[letter].width = float(width)
            except (TypeError, ValueError):
                continue

    # Row heights.
    rh = sheet_def.get("row_heights") or {}
    if isinstance(rh, dict):
        for r_str, height in rh.items():
            try:
                ws.row_dimensions[int(r_str)].height = float(height)
            except (TypeError, ValueError):
                continue

    # Merges.
    merges = sheet_def.get("merges") or []
    if isinstance(merges, list):
        for rng in merges:
            if isinstance(rng, str):
                try:
                    ws.merge_cells(rng)
                except (ValueError, TypeError):
                    continue

    # Freeze panes.
    freeze = sheet_def.get("freeze_panes")
    if isinstance(freeze, str):
        ws.freeze_panes = freeze


# ---------------------------------------------------------------------------
# Cell prop translation (mirrors xlsx_ops handlers; kept in sync manually)
# ---------------------------------------------------------------------------


def _apply_legacy_cell_props(cell, props: dict) -> None:
    if "value" in props:
        cell.value = _coerce_value(props["value"])
    if "number_format" in props and isinstance(props["number_format"], str):
        cell.number_format = props["number_format"]
    if isinstance(props.get("font"), dict):
        kw = {k: props["font"][k] for k in
              ("name", "size", "bold", "italic", "underline", "strike", "color")
              if k in props["font"]}
        if kw:
            cell.font = Font(**kw)
    if isinstance(props.get("fill"), dict):
        kw = {k: props["fill"][k] for k in
              ("fill_type", "start_color", "end_color")
              if k in props["fill"]}
        if "start_color" in kw and "fill_type" not in kw:
            kw["fill_type"] = "solid"
        if kw:
            cell.fill = PatternFill(**kw)
    if isinstance(props.get("alignment"), dict):
        kw = {k: props["alignment"][k] for k in
              ("horizontal", "vertical", "wrap_text", "shrink_to_fit",
               "indent", "text_rotation")
              if k in props["alignment"]}
        if kw:
            cell.alignment = Alignment(**kw)
    if isinstance(props.get("border"), dict):
        sides: dict = {}
        for side_name in ("left", "right", "top", "bottom", "diagonal"):
            side_def = props["border"].get(side_name)
            if isinstance(side_def, dict):
                sides[side_name] = Side(**{
                    k: side_def[k] for k in ("style", "color") if k in side_def
                })
        if sides:
            cell.border = Border(**sides)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _col_label(c: Any) -> str:
    if isinstance(c, str):
        return c
    if isinstance(c, dict):
        return str(c.get("label") or c.get("name") or c.get("id") or "")
    return str(c) if c is not None else ""


def _coerce_value(v: Any) -> Any:
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, (list, dict)):
        import json as _json
        return _json.dumps(v, ensure_ascii=False)
    return v
